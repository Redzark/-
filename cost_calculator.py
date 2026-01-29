import streamlit as st
import openpyxl
import io
import uuid
import re
import traceback
import zipfile
from copy import copy
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# ============================================================================
# 1. ê¸°ì´ˆ ë°ì´í„° ë° ì„¤ì • (ì‚¬ì¥ë‹˜ ê¸°ì¤€ ì ˆëŒ€ ì‚¬ìˆ˜)
# ============================================================================
MAT_START_ROW = 12
MAT_STEP = 4
LAB_START_ROW = 55
EXP_START_ROW = 75
TEMPLATE_HEIGHT = 85 

YEARLY_LABOR_RATES = {2025: 25700, 2026: 30000, 2027: 31600, 2028: 33200, 2029: 34800}
DIRECT_EXP_TABLE = {50: 2042, 70: 2248, 100: 2735, 120: 2819, 150: 3230, 170: 3219, 220: 4404, 250: 4861, 300: 6210, 350: 6349, 450: 8204, 500: 7940, 550: 9228, 600: 10009, 650: 11482, 700: 11488, 750: 13458, 850: 14604, 900: 15154, 1050: 17575, 1300: 21270, 1600: 23872, 1800: 27671, 2000: 27671, 2200: 33488, 2300: 33488, 2400: 33488, 2500: 33488, 3000: 48003}
MATERIAL_DATA = {
    "ë¬´ë„ì¥ TPO": {"coeff": 2.58, "f12": "ë¬´ë„ì¥ TPO", "f13": "MS220-19 TYPE B-2"},
    "ë„ì¥ìš© TPO": {"coeff": 3.56, "f12": "ë„ì¥ìš© TPO", "f13": "MS220-19 TYPE B-1"},
    "ASA": {"coeff": 3.11, "f12": "ASA-022 TYPE B", "f13": "MS225-22"},
    "ë„ê¸ˆìš© ABS": {"coeff": 3.62, "f12": "ë„ê¸ˆìš© ABS", "f13": "MS225-20"},
    "ë„ì¥ìš© ABS": {"coeff": 3.62, "f12": "ë„ì¥ìš© ABS", "f13": "MS225-18 TYPE C"},
    "PP": {"coeff": 2.58, "f12": "PP", "f13": "General"}
}
DRY_CYCLE_MAP = {50:10, 70:11, 100:12, 120:13, 150:14, 170:14, 220:15, 280:16, 350:19, 450:21, 500:21, 550:21, 600:22, 650:22, 700:23, 750:23, 850:26, 900:26, 1050:26, 1300:28, 1600:30, 1800:31, 2000:32, 2200:36, 2300:37, 2400:37, 2500:38, 3000:44}

# ============================================================================
# 2. ë¡œì§ í•¨ìˆ˜
# ============================================================================
def safe_float(value, default=0.0):
    try:
        if value is None: return default
        s_val = str(value).strip().upper()
        if not s_val: return default
        for sep in ['\n', '(', '\r']:
            if sep in s_val: s_val = s_val.split(sep)[0].strip()
        if "/" in s_val:
            parts = s_val.split("/")
            if parts[0].strip(): s_val = parts[0]
        clean_val = re.sub(r"[^0-9.]", "", s_val)
        if not clean_val: return default
        if clean_val == ".": return default
        return float(clean_val)
    except: return default

def get_loss_rate(real_vol):
    if real_vol <= 3000: return 0.049
    elif real_vol <= 5000: return 0.032
    elif real_vol <= 10000: return 0.019
    elif real_vol <= 20000: return 0.010
    elif real_vol <= 40000: return 0.006
    elif real_vol <= 80000: return 0.006
    elif real_vol <= 100000: return 0.005
    elif real_vol <= 200000: return 0.005
    elif real_vol <= 300000: return 0.003
    elif real_vol <= 400000: return 0.002
    elif real_vol <= 600000: return 0.002
    elif real_vol <= 800000: return 0.002
    else: return 0.001 

def get_lot_size(L, W, H, real_vol):
    max_dim = max(L, W, H)
    idx = 1 if max_dim <= 100 else (3 if max_dim > 1500 else 2)
    return 5000 if idx==1 else (3000 if idx==2 else 1500)

def get_manpower(ton, mat_name):
    if "ë„ê¸ˆ" in mat_name: return 0.5 if ton <= 150 else 1.0
    return 0.5 if ton < 650 else 1.0

def get_setup_time(ton):
    if ton <= 150: return 25
    elif ton < 650: return 30
    elif ton < 2000: return 35
    else: return 45

def get_sr_rate_value(w, c):
    total = w * c
    if total <= 3: return 240
    elif total <= 5: return 160
    elif total <= 10: return 90
    elif total <= 30: return 35
    elif total <= 50: return 22
    elif total <= 200: return 8
    elif total <= 1500: return 5
    elif total <= 3000: return 4
    else: return 3

def get_machine_factor(ton):
    if ton < 150: return 0.9
    elif ton < 300: return 1.0
    elif ton < 650: return 1.05
    elif ton < 1300: return 1.1
    elif ton < 1800: return 1.2
    else: return 1.3

def get_depth_factor(h):
    if h <= 50: return 0.8
    elif h <= 100: return 0.9
    elif h <= 150: return 0.95
    elif h <= 200: return 1.0
    elif h <= 250: return 1.05
    elif h <= 300: return 1.1
    elif h <= 400: return 1.15
    else: return 1.2

def safe_write(ws, coord, value):
    try: ws[coord] = value
    except Exception: pass

# ============================================================================
# 3. PART LIST íŒŒì‹± í•¨ìˆ˜ (ë¸”ë¡ ë‹¨ìœ„ íŒŒì‹±: ì™„ë²½ ë¶„í•´)
# ============================================================================
def normalize_header(s):
    if not s: return ""
    return re.sub(r'[^A-Z0-9]', '', str(s).upper())

def extract_header_info(ws):
    extracted = {"car": "", "vol": 0}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=150, values_only=True)):
        for j, cell in enumerate(row):
            if not cell: continue
            s_val = normalize_header(cell)
            if "PROJECT" in s_val or "ì°¨ì¢…" in s_val:
                for k in range(j + 1, len(row)):
                    if row[k]: extracted["car"] = str(row[k]).strip(); break
            if "VOLUME" in s_val or "ìƒì‚°ëŒ€ìˆ˜" in s_val or "ìƒì‚°ëŸ‰" in s_val:
                for k in range(j, min(j + 10, len(row))): 
                    val = safe_float(row[k])
                    if val > 0: extracted["vol"] = val; break 
    return extracted

def parse_part_list_matrix(file):
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        header_info = extract_header_info(ws)
        all_rows = list(ws.iter_rows(values_only=True))
        
        # 1. í—¤ë” íƒìƒ‰
        header_row_index = -1
        col_map = {'lv_start': 1, 'part_no': -1, 'name': -1, 'qty_cols': [], 'mat': -1, 'ton': -1, 'cav': -1, 'L': -1, 'W': -1, 'H': -1}
        
        debug_log = []

        for i in range(min(25, len(all_rows))):
            r = all_rows[i]
            row_norm = "".join([normalize_header(x) for x in r])
            if "PARTNO" in row_norm or "í’ˆë²ˆ" in row_norm:
                header_row_index = i
                break
        
        if header_row_index == -1: header_row_index = 5

        rows_to_scan = [all_rows[header_row_index]]
        if header_row_index + 1 < len(all_rows): rows_to_scan.append(all_rows[header_row_index + 1])

        qty_cols_set = set()

        for r_idx, r in enumerate(rows_to_scan):
            for idx, cell in enumerate(r):
                if not cell: continue
                s_val = normalize_header(cell)
                if "LV" == s_val or "LEVEL" in s_val: col_map['lv_start'] = idx
                elif "PARTNO" in s_val or "í’ˆë²ˆ" in s_val: col_map['part_no'] = idx
                elif "PARTNAME" in s_val or "í’ˆëª…" in s_val: col_map['name'] = idx
                elif "MATERIAL" in s_val or "ì¬ì§ˆ" in s_val: col_map['mat'] = idx
                elif "THICK" in s_val or "ë‘ê»˜" in s_val: col_map['thick'] = idx
                elif "WEIGHT" in s_val or "ì¤‘ëŸ‰" in s_val: col_map['weight'] = idx
                elif "TON" in s_val or "í†¤" in s_val: col_map['ton'] = idx
                elif "CAV" in s_val: col_map['cav'] = idx
                elif s_val in ["L", "LENGTH", "ê°€ë¡œ"]: col_map['L'] = idx
                elif s_val in ["W", "WIDTH", "ì„¸ë¡œ"]: col_map['W'] = idx
                elif s_val in ["H", "HEIGHT", "ë†’ì´", "ê¹Šì´"]: col_map['H'] = idx
                
                # ìˆ˜ëŸ‰ ì»¬ëŸ¼ ê°ì§€
                if "QTY" in s_val or "USG" in s_val or "USAGE" in s_val or "ìˆ˜ëŸ‰" in s_val:
                    qty_cols_set.add(idx)

        col_map['qty_cols'] = sorted(list(qty_cols_set))
        if col_map['lv_start'] == -1: col_map['lv_start'] = 1 

        debug_log.append(f"â„¹ï¸ ìˆ˜ëŸ‰ ê¸°ë‘¥ {len(col_map['qty_cols'])}ê°œ ê°ì§€ë¨")

        # 2. íŒŒì‹± (ë¸”ë¡ ë‹¨ìœ„ - Block Parsing)
        assy_dict = {} 
        
        lv_start = col_map['lv_start']
        lv_end = col_map['part_no'] if col_map['part_no'] != -1 else lv_start + 5

        # Lv.1 í–‰ë“¤ì˜ ì¸ë±ìŠ¤ë¥¼ ë¨¼ì € ì°¾ìŠµë‹ˆë‹¤.
        level1_indices = []
        for i in range(header_row_index + 1, len(all_rows)):
            r = list(all_rows[i])
            # Lv.1 ì—¬ë¶€ í™•ì¸
            this_level = 999
            for l_idx in range(lv_start, lv_end + 2):
                if l_idx >= len(r): break
                val = str(r[l_idx]).strip()
                if "â—" in val or "1" == val or "â—" in normalize_header(val):
                    this_level = l_idx - lv_start + 1
                    break
            if this_level == 1:
                level1_indices.append(i)
        
        level1_indices.append(len(all_rows)) # ë ì§€ì  ì¶”ê°€

        # ê° Lv.1 ë¸”ë¡ì„ ìˆœíšŒ
        for k in range(len(level1_indices) - 1):
            start_row = level1_indices[k]
            end_row = level1_indices[k+1]
            
            # Lv.1 í–‰ ë°ì´í„°
            root_r = list(all_rows[start_row])
            if len(root_r) < 100: root_r.extend([None] * (100 - len(root_r)))

            # ì´ Lv.1ì´ ì–´ë–¤ ê¸°ë‘¥(Column)ì—ì„œ í™œì„±í™”ë˜ì–´ ìˆëŠ”ì§€ í™•ì¸
            # ì˜ˆ: 86350-T6700ì€ Jì—´ì—ì„œ í™œì„±, Kì—´ì—ì„  ë¹„í™œì„±
            
            p_idx = col_map.get('part_no', 7)
            n_idx = col_map.get('name', 8)
            raw_no = str(root_r[p_idx]).strip() if (p_idx != -1 and root_r[p_idx]) else ""
            raw_name = str(root_r[n_idx]).strip() if (n_idx != -1 and root_r[n_idx]) else f"ASSY_{k}"
            
            if not raw_no or "ASSY" in raw_no.upper() or "í•„ìš”" in raw_no:
                base_name_origin = raw_name.replace("/", "_").replace("*", "")[:35]
            else:
                base_name_origin = raw_no.replace("/", "_").replace("*", "")

            # ì‚¬ìš©ëœ ê¸°ë‘¥ ì°¾ê¸°
            active_cols = []
            for q_col in col_map['qty_cols']:
                if safe_float(root_r[q_col]) > 0:
                    active_cols.append(q_col)
            
            if not active_cols:
                # Lv.1ì¸ë° ìˆ˜ëŸ‰ì´ í•˜ë‚˜ë„ ì—†ëŠ” ê²½ìš° -> ê·¸ë˜ë„ í•˜ìœ„ ë¶€í’ˆì€ ìˆì„ ìˆ˜ ìˆìœ¼ë‹ˆ ì¼ë‹¨ ì§„í–‰?
                # ë³´í†µ ì´ëŸ° ê²½ìš° Header ì—­í• ë§Œ í•¨.
                # í•˜ì§€ë§Œ í•˜ìœ„ ë¶€í’ˆì´ íŠ¹ì • ì—´ì— ìˆ˜ëŸ‰ì´ ìˆë‹¤ë©´ ê·¸ ì—´ì˜ ASSYë¡œ ê°„ì£¼í•´ì•¼ í•¨.
                # í¸ì˜ìƒ ëª¨ë“  ìˆ˜ëŸ‰ ê¸°ë‘¥ì„ ê²€ì‚¬ ëŒ€ìƒìœ¼ë¡œ ì‚¼ìŒ.
                active_cols = col_map['qty_cols']

            # ì´ì œ ì´ ë¸”ë¡(start_row ~ end_row) ì•ˆì— ìˆëŠ” ë¶€í’ˆë“¤ì„ ê¸ì–´ëª¨ìŒ
            # ë‹¨, íŒŒì¼ì€ Lv.1 ì´ë¦„ ê¸°ì¤€ìœ¼ë¡œ ìƒì„± (ì¤‘ë³µë˜ë©´ _Col ì¶”ê°€)
            
            # ëŒ€ì¥(Lv.1)ì´ í•˜ë‚˜ë¼ë„ ë°œê²¬ë˜ë©´ íŒŒì¼ ìƒì„± ëŒ€ìƒ
            # ì—¬ê¸°ì„œëŠ” "Lv.1 í’ˆë²ˆ" ìì²´ê°€ í•˜ë‚˜ì˜ íŒŒì¼ì´ ë¨.
            
            # [ì¤‘ìš”] Lv.1ì´ ì—¬ëŸ¬ ê¸°ë‘¥ì— ê±¸ì³ ìˆì„ ìˆ˜ ìˆìŒ (ê³µìš© Back Cover ë“±)
            # ì´ ê²½ìš° íŒŒì¼ í•˜ë‚˜ë§Œ ë§Œë“¤ë©´ ë¨.
            
            assy_key = base_name_origin
            if assy_key in assy_dict: 
                # ì´ë¦„ì´ ì¤‘ë³µëœë‹¤ë©´ (ë“œë¬¸ ê²½ìš°), ë’¤ì— ë²ˆí˜¸ ë¶™ì„
                assy_key = f"{assy_key}_{k}"
            
            items_in_this_block = []

            # Lv.1 ë³¸ì¸ì´ ì‚¬ì¶œí’ˆì´ë©´ ì¶”ê°€
            # ... (ë³¸ì¸ ì¶”ê°€ ë¡œì§ ìƒëµ, ë³´í†µ Assemblyë¼ ì œì™¸í•˜ì§€ë§Œ í†¤ìˆ˜ ìˆìœ¼ë©´ ì¶”ê°€)
            t_idx = col_map.get('ton', 28)
            m_idx = col_map.get('mat', 27) 
            raw_ton = root_r[t_idx] if t_idx != -1 and t_idx < len(root_r) else None
            raw_mat = root_r[m_idx] if m_idx != -1 and m_idx < len(root_r) else None
            
            if safe_float(raw_ton) or (raw_mat and str(raw_mat).strip()):
                 # ë³¸ì¸ ì¶”ê°€
                 # (ì¶”ê°€ ì½”ë“œëŠ” ì•„ë˜ í•˜ìœ„ ë¶€í’ˆ ë¡œì§ê³¼ ë™ì¼í•˜ë¯€ë¡œ í•¨ìˆ˜í™”í•˜ê±°ë‚˜ ë³µë¶™)
                 pass # ì¼ë‹¨ ìƒëµ, ë³´í†µ í•˜ìœ„ë¶€í’ˆì´ ì¤‘ìš”

            # í•˜ìœ„ ë¶€í’ˆ ìˆœíšŒ (start_row + 1 ~ end_row - 1)
            for i in range(start_row + 1, end_row):
                r = list(all_rows[i])
                if len(r) < 100: r.extend([None] * (100 - len(r)))
                
                # ì´ ë¶€í’ˆì´ ì´ Lv.1 ë¸”ë¡ì—ì„œ ìœ íš¨í•œì§€ í™•ì¸
                # ì¡°ê±´: Lv.1ì´ í™œì„±í™”ëœ ê¸°ë‘¥(active_cols) ì¤‘ í•˜ë‚˜ë¼ë„ ìˆ˜ëŸ‰ì´ ìˆì–´ì•¼ í•¨.
                
                is_valid_child = False
                usage_val = 0.0
                
                for q_col in active_cols:
                    u = safe_float(r[q_col])
                    if u > 0:
                        is_valid_child = True
                        usage_val = u # ì²« ë²ˆì§¸ ë°œê²¬ëœ ìˆ˜ëŸ‰ ì‚¬ìš© (ë³´í†µ ê°™ìŒ)
                        break
                
                if is_valid_child:
                    # ì‚¬ì¶œí’ˆ ì¡°ê±´ í™•ì¸
                    raw_ton = r[t_idx] if t_idx != -1 and t_idx < len(r) else None
                    raw_mat = r[m_idx] if m_idx != -1 and m_idx < len(r) else None
                    
                    if safe_float(raw_ton) or (raw_mat and str(raw_mat).strip()):
                        # ë°ì´í„° ì¶”ì¶œ
                        p_idx = col_map['part_no']
                        n_idx = col_map['name']
                        p_no_str = str(r[p_idx]).strip() if (p_idx != -1 and r[p_idx]) else ""
                        p_name = str(r[n_idx]).strip() if (n_idx != -1 and r[n_idx]) else ""
                        
                        l = safe_float(r[col_map['L']]) if col_map['L'] != -1 else 0
                        w = safe_float(r[col_map['W']]) if col_map['W'] != -1 else 0
                        h = safe_float(r[col_map['H']]) if col_map['H'] != -1 else 0
                        t_col = col_map['thick']
                        t = safe_float(r[t_col]) if (t_col != -1 and t_col < len(r)) else 2.5
                        if t == 0: t = 2.5
                        w_col = col_map['weight']
                        weight_val = safe_float(r[w_col]) if (w_col != -1 and w_col < len(r)) else 0.0

                        mapped_mat = "ë¬´ë„ì¥ TPO"
                        if raw_mat:
                            s_mat = str(raw_mat).upper()
                            for key in MATERIAL_DATA.keys():
                                if key in s_mat: mapped_mat = key; break
                            if "PP" in s_mat and mapped_mat == "ë¬´ë„ì¥ TPO": mapped_mat = "PP"

                        ton = int(safe_float(raw_ton, default=1300))
                        
                        cv_idx = col_map['cav']
                        raw_cav = str(r[cv_idx]) if (cv_idx != -1 and cv_idx < len(r)) else "1"
                        if "/" in raw_cav:
                            try: cav = int(sum(safe_float(x) for x in raw_cav.split('/') if x.strip()))
                            except: cav = int(safe_float(raw_cav, default=1))
                        else:
                            cav = int(safe_float(raw_cav, default=1))
                        if cav < 1: cav = 1

                        item = {
                            "id": str(uuid.uuid4()),
                            "level": "ì‚¬ì¶œì œí’ˆ",
                            "no": p_no_str,
                            "name": p_name,
                            "remarks": str(r[n_idx + 1] if n_idx != -1 and n_idx + 1 < len(r) and r[n_idx+1] else ""),
                            "opt_rate": 100.0,
                            "usage": usage_val, 
                            "L": l, "W": w, "H": h, "thick": t,
                            "weight": weight_val,
                            "mat": mapped_mat,
                            "ton": ton,
                            "cavity": cav,
                            "price": 2000
                        }
                        items_in_this_block.append(item)
            
            if items_in_this_block:
                assy_dict[assy_key] = items_in_this_block
                debug_log.append(f"ğŸ“¦ ASSY ìƒì„±: {assy_key} (ë¶€í’ˆ {len(items_in_this_block)}ê°œ)")

        return assy_dict, header_info, debug_log

    except Exception as e:
        return {}, {}, [f"âŒ ì˜¤ë¥˜: {str(e)}", traceback.format_exc()]

# ============================================================================
# 4. ì—‘ì…€ ìƒì„± í•¨ìˆ˜ (ìˆ˜ì§ ì´ì–´ë¶™ì´ê¸°)
# ============================================================================
def copy_template_style(src_ws, tgt_ws, start_row, max_row):
    for row in range(1, max_row + 1):
        for col in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=row, column=col)
            tgt_cell = tgt_ws.cell(row=start_row + row - 1, column=col)
            tgt_cell.value = src_cell.value
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = src_cell.number_format
                tgt_cell.protection = copy(src_cell.protection)
                tgt_cell.alignment = copy(src_cell.alignment)
    for merged_cell in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell.bounds
        tgt_ws.merge_cells(start_row=start_row + min_row - 1, start_column=min_col, end_row=start_row + max_row - 1, end_column=max_col)
    if start_row == 1:
        for i in range(1, src_ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            if col_letter in src_ws.column_dimensions:
                tgt_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width

def generate_excel_file_stacked(common, items, sel_year):
    try:
        wb = openpyxl.load_workbook("template.xlsx")
        template_ws = wb.active 
    except: return None

    # [1] ì§‘ê³„í‘œ ì‹œíŠ¸
    ws_summary = wb.create_sheet("ASSY_Summary", 0)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="36486b", end_color="36486b", fill_type="solid")
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["NO", "PART NO", "PART NAME", "USAGE", "MATERIAL", "TON", "CAVITY", "WEIGHT(g)", "NOTE"]
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    
    for idx, item in enumerate(items, 1):
        row_num = idx + 1
        data = [idx, item['no'], item['name'], item['usage'], item['mat'], item['ton'], item['cavity'], item['weight'], item['remarks']]
        for col_idx, val in enumerate(data, 1):
            cell = ws_summary.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = align_center
            cell.border = thin_border
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 35

    # [2] ìƒì„¸ ì‹œíŠ¸
    ws_main = wb.create_sheet("Calculation_Total", 1)
    template_max_row = template_ws.max_row
    current_offset = 1 

    for item in items:
        copy_template_style(template_ws, ws_main, current_offset, template_max_row)
        
        def cell(r): 
            col = openpyxl.utils.cell.coordinate_from_string(r)[0]
            row = openpyxl.utils.cell.coordinate_from_string(r)[1]
            return f"{col}{row + current_offset - 1}"

        safe_write(ws_main, cell("N3"), common['car'])
        safe_write(ws_main, cell("C3"), item['no'])    
        safe_write(ws_main, cell("C4"), item['name']) 
        
        def r_idx(base_row): return base_row + current_offset - 1
        curr_m = r_idx(MAT_START_ROW)
        
        item_usage = item.get('usage', 1.0)
        real_vol = common['base_vol'] * (item['opt_rate'] / 100) * item_usage
        loss_val = get_loss_rate(real_vol)

        safe_write(ws_main, f"B{curr_m}", item['name'])
        safe_write(ws_main, f"B{curr_m+1}", item['no'])
        
        mat_info = MATERIAL_DATA.get(item['mat'], MATERIAL_DATA["ë¬´ë„ì¥ TPO"])
        safe_write(ws_main, f"F{curr_m}", mat_info['f12'])
        safe_write(ws_main, f"F{curr_m+1}", mat_info['f13'])
        if ws_main[f"F{curr_m}"]: ws_main[f"F{curr_m}"].alignment = align_center
        if ws_main[f"F{curr_m+1}"]: ws_main[f"F{curr_m+1}"].alignment = align_center
        
        safe_write(ws_main, f"D{curr_m}", real_vol) 
        if ws_main[f"D{curr_m}"]: ws_main[f"D{curr_m}"].number_format = '#,##0'
        
        safe_write(ws_main, f"J{curr_m}", item['weight']/1000)
        safe_write(ws_main, f"K{curr_m}", item['price'])
        safe_write(ws_main, f"H{curr_m}", 1.0)
        safe_write(ws_main, f"I{curr_m}", "kg")
        if ws_main[f"I{curr_m}"]: ws_main[f"I{curr_m}"].alignment = align_center
        safe_write(ws_main, f"L{curr_m}", f"=(J{curr_m}*(1+{loss_val}))*K{curr_m}*H{curr_m}")
        
        sr_val = get_sr_rate_value(item['weight'], item['cavity'])
        safe_write(ws_main, f"J{curr_m+1}", f"=J{curr_m} * {sr_val} / 100")
        safe_write(ws_main, f"K{curr_m+1}", 87)
        safe_write(ws_main, f"H{curr_m+1}", 1.0)
        safe_write(ws_main, f"I{curr_m+1}", "kg")
        if ws_main[f"I{curr_m+1}"]: ws_main[f"I{curr_m+1}"].alignment = align_center
        safe_write(ws_main, f"L{curr_m+1}", f"=J{curr_m+1}*K{curr_m+1}*H{curr_m+1}")

        l_row = r_idx(LAB_START_ROW)
        e_row = r_idx(EXP_START_ROW)

        setup, lot = get_setup_time(item['ton']), get_lot_size(item['L'], item['W'], item['H'], real_vol)
        mp, l_rate, e_rate = get_manpower(item['ton'], item['mat']), YEARLY_LABOR_RATES[sel_year], DIRECT_EXP_TABLE.get(item['ton'], 7940)
        
        safe_write(ws_main, f"B{l_row}", item['name'])
        safe_write(ws_main, f"F{l_row}", setup)
        if ws_main[f"F{l_row}"]: ws_main[f"F{l_row}"].alignment = align_center
        safe_write(ws_main, f"G{l_row}", lot)
        if ws_main[f"G{l_row}"]: ws_main[f"G{l_row}"].alignment = align_center
        safe_write(ws_main, f"H{l_row}", item['cavity'])
        safe_write(ws_main, f"I{l_row}", mp)
        safe_write(ws_main, f"K{l_row}", l_rate)
        safe_write(ws_main, f"E{l_row}", 1.0)

        mf, hf, dry = get_machine_factor(item['ton']), get_depth_factor(item.get('H', 100)), DRY_CYCLE_MAP.get(item['ton'], 44)
        ct_formula = f"={dry}+(4.396*((SUM(J{curr_m}:J{curr_m+1})*H{l_row})*1000)^0.1477)+({MATERIAL_DATA.get(item['mat'], MATERIAL_DATA['ë¬´ë„ì¥ TPO'])['coeff']}*{item.get('thick', 2.5)}^2*{mf}*{hf})"
        if item['mat'] == "ë„ê¸ˆìš© ABS": ct_formula += "+15"
        safe_write(ws_main, f"J{l_row}", ct_formula)
        safe_write(ws_main, f"L{l_row}", f"=(J{l_row}*1.1/H{l_row}+F{l_row}*60/G{l_row})*I{l_row}*K{l_row}/3600*E{l_row}") 

        safe_write(ws_main, f"B{e_row}", item['name'])
        safe_write(ws_main, f"F{e_row}", setup)
        safe_write(ws_main, f"G{e_row}", lot)
        safe_write(ws_main, f"H{e_row}", item['cavity'])
        safe_write(ws_main, f"I{e_row}", item['ton'])
        if ws_main[f"I{e_row}"]: ws_main[f"I{e_row}"].number_format = '#,##0"T"'
        safe_write(ws_main, f"J{e_row}", f"=J{l_row}")
        safe_write(ws_main, f"K{e_row}", e_rate)
        safe_write(ws_main, f"E{e_row}", 1.0) 
        safe_write(ws_main, f"L{e_row}", f"=(J{l_row}*1.1/H{e_row}+F{e_row}*60/G{e_row})*K{e_row}/3600*(1+0.64)")
        
        current_offset += (template_max_row + 2)

    if "Master_Template" in wb.sheetnames: wb.remove(wb["Master_Template"])
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ============================================================================
# 5. Streamlit UI
# ============================================================================
st.set_page_config(page_title="ì›ê°€ê³„ì‚°ì„œ(í†µí•©)", layout="wide")
st.title("ì›ê°€ê³„ì‚°ì„œ (ë‹¨í’ˆ/ìˆ˜ë™ + ASSY í†µí•©ë³¸)")

if 'manual_items' not in st.session_state: st.session_state.manual_items = []
if 'assy_dict' not in st.session_state: st.session_state.assy_dict = {}
if 'common_car' not in st.session_state: st.session_state.common_car = ""
if 'common_vol' not in st.session_state: st.session_state.common_vol = 0

mode = st.radio("ì‘ì—… ëª¨ë“œ ì„ íƒ", ["ë‹¨í’ˆ ê³„ì‚°", "ASSY(ìˆ˜ë™ ì…ë ¥)", "PART LIST ì—‘ì…€ ì—…ë¡œë“œ(Matrix)"], horizontal=True)

if mode in ["ë‹¨í’ˆ ê³„ì‚°", "ASSY(ìˆ˜ë™ ì…ë ¥)"]:
    st.info("ğŸ’¡ ì§ì ‘ ë°ì´í„°ë¥¼ ì…ë ¥í•˜ì—¬ ê³„ì‚°ì„œë¥¼ ë§Œë“­ë‹ˆë‹¤.")
    c1, c2, c3 = st.columns(3)
    car = c1.text_input("ì°¨ì¢…", value=st.session_state.common_car)
    base_vol = c2.number_input("ê¸°ë³¸ Volume (ëŒ€)", value=int(st.session_state.common_vol) if st.session_state.common_vol else 0)

    if mode == "ë‹¨í’ˆ ê³„ì‚°" and not st.session_state.manual_items:
        st.session_state.manual_items = [{"id":str(uuid.uuid4()), "level":"ì‚¬ì¶œì œí’ˆ", "no":"", "name":"", "opt_rate":100.0, "usage":1.0, "L":0.0, "W":0.0, "H":0.0, "thick":2.5, "weight":0.0, "mat":"ë¬´ë„ì¥ TPO", "ton":1300, "cavity":1, "price":2000}]
    
    if mode == "ASSY(ìˆ˜ë™ ì…ë ¥)":
        if st.button("â• í’ˆëª© ì¶”ê°€"):
            st.session_state.manual_items.append({"id":str(uuid.uuid4()), "level":"ì‚¬ì¶œì œí’ˆ", "no":"", "name":"", "opt_rate":100.0, "usage":1.0, "L":0.0, "W":0.0, "H":0.0, "thick":2.5, "weight":0.0, "mat":"ë¬´ë„ì¥ TPO", "ton":1300, "cavity":1, "price":2000})

    for i, item in enumerate(st.session_state.manual_items):
        uid = item['id']
        with st.container(border=True):
            cols = st.columns([2, 2, 2, 1, 1, 0.5])
            item['no'] = cols[0].text_input("í’ˆë²ˆ", value=item['no'], key=f"n_{uid}")
            item['name'] = cols[1].text_input("í’ˆëª…", value=item['name'], key=f"nm_{uid}")
            item['opt_rate'] = cols[2].number_input("ì˜µì…˜ìœ¨(%)", value=item['opt_rate'], key=f"op_{uid}")
            item['usage'] = cols[3].number_input("Qty", value=item['usage'], key=f"us_{uid}")
            if mode == "ASSY(ìˆ˜ë™ ì…ë ¥)":
                if cols[5].button("ğŸ—‘ï¸", key=f"d_{uid}"): st.session_state.manual_items.pop(i); st.rerun()
            r = st.columns(5)
            item['L'] = r[0].number_input("L", value=item['L'], key=f"l_{uid}")
            item['W'] = r[1].number_input("W", value=item['W'], key=f"w_{uid}")
            item['H'] = r[2].number_input("H", value=item['H'], key=f"h_{uid}")
            item['thick'] = r[3].number_input("T", value=item['thick'], key=f"t_{uid}")
            item['weight'] = r[4].number_input("ì¤‘ëŸ‰(g)", value=item['weight'], key=f"g_{uid}")
            r2 = st.columns(3)
            mat_idx = 0
            if item['mat'] in MATERIAL_DATA: mat_idx = list(MATERIAL_DATA.keys()).index(item['mat'])
            item['mat'] = r2[0].selectbox("ì†Œì¬", list(MATERIAL_DATA.keys()), index=mat_idx, key=f"ma_{uid}")
            ton_keys = list(DIRECT_EXP_TABLE.keys())
            ton_idx = ton_keys.index(item['ton']) if item['ton'] in ton_keys else ton_keys.index(1300)
            item['ton'] = r2[1].selectbox("Ton", ton_keys, index=ton_idx, key=f"to_{uid}")
            item['cavity'] = r2[2].number_input("Cav", min_value=1, value=int(item['cavity']), key=f"ca_{uid}")
            item['price'] = st.number_input("ë‹¨ê°€(ì°¸ê³ ìš©)", value=item['price'], key=f"pr_{uid}")

    if st.button("ì—‘ì…€ ìƒì„±", type="primary"):
        excel_bytes = generate_excel_file_stacked({"car":car, "base_vol":base_vol}, st.session_state.manual_items, 2026)
        if excel_bytes: st.download_button("ğŸ“¥ ë‹¤ìš´ë¡œë“œ", excel_bytes, "Manual_Cost.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("ğŸ’¡ ì—‘ì…€ì„ ì˜¬ë¦¬ë©´ [Lv.1 ë©ì–´ë¦¬] ë‹¨ìœ„ë¡œ ìë™ ë¶„ì„í•˜ì—¬ ZIPìœ¼ë¡œ ì¤ë‹ˆë‹¤.")
    uploaded_file = st.file_uploader("PART LIST íŒŒì¼ ì—…ë¡œë“œ", type=["xlsx", "xls"])
    if uploaded_file:
        if st.button("ğŸ”„ ë¶„ì„ ì‹œì‘"):
            assy_data, info, debug_log = parse_part_list_matrix(uploaded_file)
            
            with st.expander("ğŸ” ë¶„ì„ ë¦¬í¬íŠ¸ (ëˆŒëŸ¬ì„œ í™•ì¸)", expanded=True):
                for log in debug_log: st.write(log)
            
            if assy_data:
                st.session_state.assy_dict = assy_data
                st.session_state.common_car = info.get('car', '')
                st.session_state.common_vol = info.get('vol', 0)
                st.success(f"âœ… ì´ {len(assy_data)}ê°œì˜ ASSY íŒŒì¼ì´ ìƒì„±ë  ì˜ˆì •ì…ë‹ˆë‹¤!")
            else: st.error("ë°ì´í„° ì—†ìŒ (Qt'y í—¤ë” ë˜ëŠ” í†¤ìˆ˜/ì¬ì§ˆ í™•ì¸ í•„ìš”)")

    if st.session_state.assy_dict:
        c1, c2 = st.columns(2)
        car = c1.text_input("ì°¨ì¢…", value=st.session_state.common_car, key="m_car")
        base_vol = c2.number_input("ê¸°ë³¸ Volume", value=int(st.session_state.common_vol), key="m_vol")
        st.markdown("---")
        for name, items in st.session_state.assy_dict.items():
            with st.expander(f"ğŸ“¦ {name} ({len(items)} parts)"):
                for it in items: st.write(f"- {it['no']} ({it['name']})")
        if st.button("ZIP ë‹¤ìš´ë¡œë“œ (ASSYë³„ í†µí•© íŒŒì¼)", type="primary"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for name, items in st.session_state.assy_dict.items():
                    xb = generate_excel_file_stacked({"car":car, "base_vol":base_vol}, items, 2026)
                    if xb: zf.writestr(f"{name}_í†µí•©ê³„ì‚°ì„œ.xlsx", xb)
            st.download_button("ğŸ“¥ ZIP ë°›ê¸°", zip_buffer.getvalue(), "Integrated_Cost_Set.zip", "application/zip")
